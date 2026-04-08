"""
生成文萃楼实地采集模板 Excel 文件
输出: data/survey_template_wencui.xlsx
"""

import json
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────── 颜色常量 ───────────────────────────
BLUE_HDR   = "1F4E79"   # 深蓝，表头文字
BLUE_FILL  = "BDD7EE"   # 浅蓝，表头背景
YELLOW     = "FFF2CC"   # 黄色，必填列背景
GREEN      = "E2EFDA"   # 绿色，可选列背景
GRAY       = "F2F2F2"   # 灰色，说明行背景
RED_TXT    = "C00000"   # 红色文字（提示）
ORANGE     = "FCE4D6"   # 橙色，示例行背景

def thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(): return Font(name='微软雅黑', bold=True, color=BLUE_HDR, size=10)
def body_font(): return Font(name='微软雅黑', size=9)
def note_font(): return Font(name='微软雅黑', italic=True, color='595959', size=8)

def fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write_header(ws, headers, row=1):
    """写表头（字段名 + 说明 + 必填标记）"""
    for col, (name, note, req, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.font = hdr_font()
        cell.fill = fill(BLUE_FILL)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border()
        set_col_width(ws, col, width)
        # 副标题行（说明）
        note_cell = ws.cell(row=row+1, column=col, value=note)
        note_cell.font = note_font()
        note_cell.fill = fill(GRAY)
        note_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        note_cell.border = thin_border()
        # 必填标记
        if req:
            req_cell = ws.cell(row=row+2, column=col, value='★必填')
            req_cell.font = Font(name='微软雅黑', bold=True, color='C00000', size=8)
            req_cell.fill = fill(YELLOW)
            req_cell.alignment = Alignment(horizontal='center')
            req_cell.border = thin_border()
        else:
            opt_cell = ws.cell(row=row+2, column=col, value='○可选')
            opt_cell.font = Font(name='微软雅黑', color='595959', size=8)
            opt_cell.fill = fill(GREEN)
            opt_cell.alignment = Alignment(horizontal='center')
            opt_cell.border = thin_border()

def write_example(ws, values, row, col_start=1):
    """写示例行（橙色背景）"""
    for col, val in enumerate(values, col_start):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name='微软雅黑', size=9, italic=True, color='7F7F7F')
        cell.fill = fill(ORANGE)
        cell.border = thin_border()
        cell.alignment = Alignment(vertical='center', wrap_text=True)

def write_data_row(ws, values, row, col_start=1, row_fill=None):
    """写数据行"""
    for col, val in enumerate(values, col_start):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = body_font()
        cell.border = thin_border()
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if row_fill:
            cell.fill = fill(row_fill)

def freeze_and_filter(ws, freeze_cell, filter_row):
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions

# ══════════════════════════════════════════════════════════════
#  Sheet 1: 路径行走记录（核心）
# ══════════════════════════════════════════════════════════════
def build_sheet_path(wb):
    ws = wb.create_sheet("①路径行走记录")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 20

    # 封面说明
    ws.merge_cells('A1:N1')
    title = ws.cell(row=1, column=1,
        value="【文萃楼实地采集】路径行走记录表  ——  每走一步填一行，记录经过的所有节点（房间/走廊/楼梯/电梯/特殊设施）")
    title.font = Font(name='微软雅黑', bold=True, size=11, color=BLUE_HDR)
    title.fill = fill("DEEAF1")
    title.alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        # (字段名, 说明/示例, 是否必填, 列宽)
        ("采集者",      "学生姓名",                  True,  10),
        ("日期",        "YYYY-MM-DD",               True,  11),
        ("路线编号",    "如 R01、R02（同一条路写同号）", True,  10),
        ("步骤序号",    "本条路线的第几步",            True,   8),
        ("当前楼层",    "1F/2F/…/10F",              True,   8),
        ("节点类型",    "房间/走廊/楼梯/电梯/卫生间/教师休息室/大厅/出入口/其他", True, 14),
        ("节点ID/房间号", "门牌号或JSON里的ID，如 C-101", True, 14),
        ("节点实际用途", "实验室/办公室/教室/机房/…",  False, 14),
        ("与上步连接方式", "直走/左转/右转/上楼/下楼/乘电梯", True, 14),
        ("步行步数",    "从上一节点到此节点的大步数",   False,  9),
        ("面朝方向",    "北/南/东/西（进入此节点时朝向）", False, 10),
        ("有无障碍",    "无/台阶/坡道/门禁/需刷卡",    False, 12),
        ("是否到达目的地", "是/否",                  False,  10),
        ("备注",        "任何异常、疑问、特殊情况",     False, 20),
    ]
    write_header(ws, headers, row=2)

    # 示例行
    ex_row = 5
    ws.cell(row=ex_row, column=1, value='【示例行，请勿修改】').font = Font(bold=True, color='C00000', size=8)
    examples = [
        "张三", "2026-04-10", "R01", 1, "1F", "出入口", "entrance_south_1F",
        "南门入口", "—", "—", "北", "无", "否", "从南门进入，直接面对走廊"
    ]
    write_example(ws, examples, ex_row, col_start=1)

    ex_row2 = 6
    examples2 = [
        "张三", "2026-04-10", "R01", 2, "1F", "走廊", "corridor_1F_C",
        "C区主走廊", "直走", 15, "北", "无", "否", "走廊宽约3m，两侧均有房间"
    ]
    write_example(ws, examples2, ex_row2, col_start=1)

    ex_row3 = 7
    examples3 = [
        "张三", "2026-04-10", "R01", 3, "1F", "房间", "C-101",
        "教室", "左转进入", 5, "西", "无", "否", "门口有门牌C101，教室门朝东"
    ]
    write_example(ws, examples3, ex_row3, col_start=1)

    ex_row4 = 8
    examples4 = [
        "张三", "2026-04-10", "R01", 4, "1F", "卫生间", "toilet_1F_C",
        "公共卫生间", "直走", 8, "北", "无", "否", "在C区走廊北端，男女各一"
    ]
    write_example(ws, examples4, ex_row4, col_start=1)

    # 留50行空白数据行
    for r in range(9, 59):
        for c in range(1, 15):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border()
            cell.fill = fill("FFFFFF")

    ws.freeze_panes = "A9"
    return ws


# ══════════════════════════════════════════════════════════════
#  Sheet 2: 走廊连通性核查
# ══════════════════════════════════════════════════════════════
def build_sheet_corridor(wb, nodes, edges):
    ws = wb.create_sheet("②走廊连通性核查")

    ws.merge_cells('A1:J1')
    title = ws.cell(row=1, column=1,
        value="【走廊连通性核查】——确认各区走廊之间能否直接步行贯通（无需上下楼）")
    title.font = Font(name='微软雅黑', bold=True, size=11, color=BLUE_HDR)
    title.fill = fill("DEEAF1")
    title.alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        ("楼层",       "1F~10F",                    True,   8),
        ("走廊A",      "JSON中的corridor ID",        True,  18),
        ("走廊B",      "JSON中的corridor ID",        True,  18),
        ("JSON现状",   "系统预设的连通关系",          False, 14),
        ("实地是否贯通","是/否/有门",                 True,  12),
        ("连接方式",   "直接贯通/有门/有台阶/不通",   True,  14),
        ("走廊宽度(m)","估算，如2.5",                False,  10),
        ("走廊长度(步)","两端距离步数",               False,  10),
        ("中间特殊节点","转角/门厅/天井等",           False,  16),
        ("备注",       "如某时段锁门、单向通行等",     False,  20),
    ]
    write_header(ws, headers, row=2)

    # 预填入JSON现有的同楼层corridor-corridor边，以及需要现场核实的关键对
    from collections import defaultdict
    nb = defaultdict(set)
    for e in edges:
        nb[e[0]].add(e[1])
        nb[e[1]].add(e[0])

    key_pairs = []
    # 已有边
    for e in edges:
        a, b = e[0], e[1]
        if 'corridor' in a and 'corridor' in b and 'inner' not in a and 'inner' not in b:
            for fl in ['1F','2F','3F','4F','5F','6F','7F','8F','9F','10F']:
                if fl in a and fl in b:
                    key_pairs.append((fl, a, b, '已连接（需确认）', '', '', '', '', ''))
                    break

    # 缺失的关键边（C-F 3F-7F）
    for fl in ['3F','4F','5F','6F','7F']:
        a = f'corridor_{fl}_C'
        b = f'corridor_{fl}_F'
        key_pairs.append((fl, a, b, '❌ 当前未连接，请核查', '', '', '', '', ''))

    # circular_hall
    key_pairs.append(('1F', 'circular_hall_1F', 'corridor_1F_G', '❌ 未连接', '', '', '', '', ''))
    key_pairs.append(('1F', 'circular_hall_1F', 'corridor_1F_H', '❌ 未连接', '', '', '', '', ''))
    key_pairs.append(('2F', 'circular_hall_2F', 'corridor_2F_G', '❌ 未连接', '', '', '', '', ''))

    row = 5
    for pair in key_pairs:
        fl, a, b, status = pair[0], pair[1], pair[2], pair[3]
        rf = "FFF2CC" if '未连接' in status else "FFFFFF"
        write_data_row(ws, [fl, a, b, status, '', '', '', '', '', ''], row, row_fill=rf)
        row += 1

    # 多留空行
    for r in range(row, row+20):
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = thin_border()

    ws.freeze_panes = "A5"
    return ws


# ══════════════════════════════════════════════════════════════
#  Sheet 3: 楼梯电梯核查
# ══════════════════════════════════════════════════════════════
def build_sheet_stairs(wb, nodes, edges):
    ws = wb.create_sheet("③楼梯电梯核查")

    ws.merge_cells('A1:K1')
    title = ws.cell(row=1, column=1,
        value="【楼梯/电梯核查】——确认每个楼梯实际可达楼层、单向/双向、是否有限制")
    title.font = Font(name='微软雅黑', bold=True, size=11, color=BLUE_HDR)
    title.fill = fill("DEEAF1")
    title.alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        ("类型",        "楼梯/电梯",                True,   8),
        ("JSON编号",    "stairs_C_1 / elevator_L 等", True, 16),
        ("位置描述",    "在几楼/哪个区/走廊哪头",     True,  18),
        ("JSON预设可达楼层", "系统设定",             False,  16),
        ("实地可达楼层", "如 1F,2F,3F",             True,  14),
        ("是否双向",    "是（上下均可）/仅上/仅下",   True,  12),
        ("限制条件",    "无/教工专用/学生禁入/门禁",  True,  12),
        ("电梯额定人数","仅电梯填写，如10人",         False, 10),
        ("标识/门牌",   "电梯/楼梯编号牌内容",        False, 12),
        ("紧急出口",    "是/否",                     False,  8),
        ("备注",        "任何特殊情况",               False, 20),
    ]
    write_header(ws, headers, row=2)

    # 预填楼梯/电梯组
    from collections import defaultdict
    nb = defaultdict(set)
    for e in edges:
        nb[e[0]].add(e[1])
        nb[e[1]].add(e[0])

    # 汇总每组可达楼层
    groups = {}
    for nid in nodes:
        if 'stair' in nid or 'elevator' in nid:
            parts = nid.rsplit('_', 1)
            if len(parts) == 2 and parts[1][:-1].isdigit():
                grp = parts[0]
                fl = parts[1]
                groups.setdefault(grp, []).append(fl)

    row = 5
    for grp in sorted(groups.keys()):
        floors = sorted(groups[grp], key=lambda x: int(x[:-1]))
        tp = '电梯' if 'elevator' in grp else '楼梯'
        floors_str = ','.join(floors)
        write_data_row(ws, [tp, grp, '', floors_str, '', '是', '无', '', '', '否', ''], row)
        row += 1

    for r in range(row, row+10):
        for c in range(1, 12):
            ws.cell(row=r, column=c).border = thin_border()

    ws.freeze_panes = "A5"
    return ws


# ══════════════════════════════════════════════════════════════
#  Sheet 4: 新增节点（卫生间/教师休息室/设施）
# ══════════════════════════════════════════════════════════════
def build_sheet_facilities(wb):
    ws = wb.create_sheet("④新增设施节点")

    ws.merge_cells('A1:L1')
    title = ws.cell(row=1, column=1,
        value="【新增设施节点】——填写JSON中尚未录入的节点：卫生间、教师休息室、大厅、出入口、打印室、会议室等")
    title.font = Font(name='微软雅黑', bold=True, size=11, color=BLUE_HDR)
    title.fill = fill("DEEAF1")
    title.alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        ("楼层",         "1F~10F",                True,   8),
        ("设施类型",     "卫生间/教师休息室/打印室/会议室/大厅/出入口/其他", True, 16),
        ("建议节点ID",   "如 toilet_1F_C（区域_楼层_类型）", True, 18),
        ("实际名称/标识","门牌或标识牌文字",        False, 16),
        ("所在走廊",     "旁边哪条走廊的corridor ID", True, 18),
        ("相对位置",     "走廊北端/南端/中间/左/右", True,  14),
        ("距走廊节点步数","约多少步",               False,  12),
        ("面朝方向",     "门朝哪个方向",             False,  10),
        ("是否可导航",   "是/否（如纯内部设施则否）", True,  10),
        ("开放时间",     "全天/工作日/特定时段",     False, 12),
        ("容纳人数",     "约值，卫生间填隔间数",      False,  10),
        ("备注",         "其他说明",                 False, 20),
    ]
    write_header(ws, headers, row=2)

    # 示例行
    examples = [
        "1F", "卫生间", "toilet_1F_C_south", "男厕/女厕", "corridor_1F_C",
        "走廊南端右侧", 5, "东", "是", "全天", "3隔间", "男女厕紧邻，先男后女"
    ]
    write_example(ws, examples, 5)

    examples2 = [
        "2F", "教师休息室", "teacher_lounge_2F_L", "教师休息室204",
        "corridor_2F_L", "走廊中段左侧", 3, "南", "否", "工作日8:00-18:00", "约10人", "需刷教工卡"
    ]
    write_example(ws, examples2, 6)

    for r in range(7, 50):
        for c in range(1, 13):
            ws.cell(row=r, column=c).border = thin_border()

    ws.freeze_panes = "A7"
    return ws


# ══════════════════════════════════════════════════════════════
#  Sheet 5: 房间门牌核对
# ══════════════════════════════════════════════════════════════
def build_sheet_rooms(wb, nodes):
    ws = wb.create_sheet("⑤房间门牌核对")

    ws.merge_cells('A1:I1')
    title = ws.cell(row=1, column=1,
        value="【房间门牌核对】——对照JSON预设ID，确认实地门牌是否一致；重点核查 pending_site_check=True 的节点")
    title.font = Font(name='微软雅黑', bold=True, size=11, color=BLUE_HDR)
    title.fill = fill("DEEAF1")
    title.alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        ("楼层",        "1F~10F",                  True,   8),
        ("JSON节点ID",  "系统中的ID",               True,  14),
        ("实际门牌号",  "现场看到的门牌（完整）",    True,  14),
        ("ID是否一致",  "是/否",                    True,  10),
        ("实际用途",    "教室/实验室/办公室/库房/…", False, 14),
        ("门朝向",      "北/南/东/西",              False,  8),
        ("是否可进入",  "可进入/常锁/已废弃",        False, 10),
        ("pending核查", "YES=JSON标记需现场确认",    False, 10),
        ("备注",        "门牌损坏/无门牌/与地图不符等", False, 20),
    ]
    write_header(ws, headers, row=2)

    # 预填 pending_site_check 节点
    pending = []
    for nid, n in nodes.items():
        if n.get('pending_site_check') or n.get('navigable') == False:
            pending.append((n.get('floor','?'), nid))

    pending.sort()
    row = 5
    for fl, nid in pending[:80]:
        write_data_row(ws, [fl, nid, '', '', '', '', '', 'YES', ''], row, row_fill="FFF2CC")
        row += 1

    # 空白行
    for r in range(row, row + 30):
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = thin_border()

    ws.freeze_panes = "A5"
    return ws


# ══════════════════════════════════════════════════════════════
#  Sheet 6: 问题反馈
# ══════════════════════════════════════════════════════════════
def build_sheet_issues(wb):
    ws = wb.create_sheet("⑥问题反馈")

    ws.merge_cells('A1:I1')
    title = ws.cell(row=1, column=1,
        value="【问题反馈】——填写发现的JSON错误、地图与实地不符、疑难情况等")
    title.font = Font(name='微软雅黑', bold=True, size=11, color=BLUE_HDR)
    title.fill = fill("DEEAF1")
    title.alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        ("楼层",         "1F~10F",                          True,   8),
        ("问题类型",     "房间号错误/走廊不通/节点缺失/坐标偏移/其他", True, 16),
        ("涉及节点ID",   "JSON中的ID（可多个，逗号分隔）",   True,  20),
        ("JSON现状描述", "系统里是什么",                    False, 20),
        ("实地实际情况", "现场看到的真实情况",               True,  24),
        ("严重程度",     "高（影响导航）/中（影响体验）/低（小问题）", True, 14),
        ("建议修改方案", "如：将X改为Y，或新增节点Z",        False, 24),
        ("采集者",       "学生姓名",                        True,  10),
        ("备注",         "",                                 False, 20),
    ]
    write_header(ws, headers, row=2)

    # 预填已知问题
    known_issues = [
        ("3F~7F", "走廊不通", "corridor_3F_C ~ corridor_7F_C",
         "3F-7F的C区走廊与其他走廊无同层连接",
         "请确认：C区走廊与F区走廊之间是否可直接步行（不需上下楼）？",
         "高", "若贯通则加边 corridor_xF_C ↔ corridor_xF_F", "", ""),
        ("1F/2F", "节点缺失", "circular_hall_1F, circular_hall_2F",
         "节点存在但无任何连接（孤立节点）",
         "请确认：1F/2F西侧中部（G区和H区之间）是否有圆形大厅/过渡门厅？如何进入？",
         "高", "根据实地情况补充连边", "", ""),
    ]
    row = 5
    for issue in known_issues:
        write_data_row(ws, list(issue), row, row_fill="FFF2CC")
        row += 1

    for r in range(row, row + 20):
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = thin_border()

    ws.freeze_panes = "A5"
    return ws


# ══════════════════════════════════════════════════════════════
#  Sheet 0: 使用说明
# ══════════════════════════════════════════════════════════════
def build_sheet_guide(wb):
    ws = wb.create_sheet("使用说明", 0)
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 80

    rows = [
        (None, "文萃楼室内导航 — 实地采集模板  v1.0"),
        (None, ""),
        ("▌", "【模板说明】"),
        ("", "本模板共 6 个工作表，分工明确，请按下方分工填写。"),
        ("", ""),
        ("▌", "【各表分工】"),
        ("①", "路径行走记录：核心表格。每走一个节点填一行，记录路径、连接方式、方向。"),
        ("②", "走廊连通性核查：确认各区走廊之间能否同层贯通，是导航拓扑的关键。"),
        ("③", "楼梯电梯核查：确认每部楼梯/电梯实际可达楼层，以及有无限制。"),
        ("④", "新增设施节点：补录卫生间、教师休息室、打印室、大厅、出入口等。"),
        ("⑤", "房间门牌核对：重点核查标黄的 pending 节点，确认门牌与JSON一致。"),
        ("⑥", "问题反馈：任何发现的错误、异常、地图与实地不符，都记录在这里。"),
        ("", ""),
        ("▌", "【采集优先级（时间紧张时按此顺序）】"),
        ("P1", "② 走廊连通性 — 直接决定导航路径正确性"),
        ("P1", "③ 楼梯电梯 — 跨层路径的关键"),
        ("P2", "① 路径记录 — 完整的行走记录，用于生成走廊中间节点"),
        ("P2", "⑥ 问题反馈 — 发现地图错误"),
        ("P3", "④ 新增设施 — 丰富POI内容"),
        ("P3", "⑤ 房间核对 — 完善数据质量"),
        ("", ""),
        ("▌", "【采集技巧】"),
        ("1.", "建议2人一组：一人行走+口述，一人填表。或行走时录音，回来再整理。"),
        ("2.", "方向识别：手机打开地图，对准楼层平面图校准后，看当前朝向即为N/S/E/W。"),
        ("3.", "步数：正常行走步幅约0.7m，可用于估算距离。"),
        ("4.", "走廊连通性最重要：每到一个走廊分叉，必须记录「往哪个方向能走通」。"),
        ("5.", "楼梯：上下各走一次，确认是否双向畅通，注意有无单向门。"),
        ("6.", "不确定的节点ID：先写描述（如「C区走廊北端右手边第一个门」），我来对应。"),
        ("", ""),
        ("▌", "【重点核查节点（已标注⚠的需要重点确认）】"),
        ("⚠", "3F~7F C区走廊：C区与F区之间是否有同层走廊相连？（当前系统认为不通）"),
        ("⚠", "1F/2F 西侧圆形大厅（circular_hall）：能否从G区或H区步行进入？"),
        ("⚠", "所有标黄的pending_site_check房间：门牌号是否与系统ID一致？"),
        ("", ""),
        ("▌", "【提交方式】"),
        ("", "填完后将xlsx文件发回，我来自动解析并更新wencui_building_v6.json。"),
    ]

    for r, (icon, text) in enumerate(rows, 1):
        ws.row_dimensions[r].height = 18
        if icon is None:
            cell = ws.cell(row=r, column=2, value=text)
            cell.font = Font(name='微软雅黑', bold=True, size=13, color=BLUE_HDR)
            cell.fill = fill("DEEAF1")
        elif icon in ("▌",):
            cell = ws.cell(row=r, column=2, value=text)
            cell.font = Font(name='微软雅黑', bold=True, size=11, color="1F4E79")
            cell.fill = fill("BDD7EE")
            ws.cell(row=r, column=1, value=icon).font = Font(bold=True, color="1F4E79")
        elif icon in ("⚠",):
            ws.cell(row=r, column=1, value=icon).font = Font(bold=True, color="C00000", size=10)
            cell = ws.cell(row=r, column=2, value=text)
            cell.font = Font(name='微软雅黑', size=9, color="C00000")
            cell.fill = fill("FFF2CC")
        else:
            ws.cell(row=r, column=1, value=icon).font = Font(name='微软雅黑', bold=True, size=9, color="595959")
            ws.cell(row=r, column=2, value=text).font = Font(name='微软雅黑', size=9)

    return ws


# ══════════════════════════════════════════════════════════════
#  主程序
# ══════════════════════════════════════════════════════════════
def main():
    import json
    with open("data/wencui_building_v6.json", encoding='utf-8') as f:
        d = json.load(f)
    nodes = d['nodes']
    edges = d['edges']

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认sheet

    build_sheet_guide(wb)
    build_sheet_path(wb)
    build_sheet_corridor(wb, nodes, edges)
    build_sheet_stairs(wb, nodes, edges)
    build_sheet_facilities(wb)
    build_sheet_rooms(wb, nodes)
    build_sheet_issues(wb)

    out = "data/survey_template_wencui.xlsx"
    wb.save(out)
    print(f"模板已生成: {out}")
    import os
    size = os.path.getsize(out)
    print(f"文件大小: {size/1024:.1f} KB")

if __name__ == "__main__":
    main()
